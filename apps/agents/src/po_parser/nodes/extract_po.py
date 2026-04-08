"""Single-node PO extraction: parse all attachments, call GPT-4o vision, normalise."""

from __future__ import annotations

import base64
import json
import logging
import os
import re
from datetime import datetime
from html import unescape
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from src.po_parser.prompts.extraction import EXTRACTION_SYSTEM_PROMPT
from src.po_parser.schemas.po import ExtractedPO, POItem
from src.po_parser.schemas.states import AgentState
from src.po_parser.tools.file_helpers import b64decode_bytes, unlink_silent, write_temp_bytes
from src.services.openai.client import OpenAIClient
from src.services.openai.settings import OpenAISettings

logger = logging.getLogger(__name__)

_openai: OpenAIClient | None = None

_HEADER_ALIASES = {
    "p.o. #": "po_number",
    "po #": "po_number",
    "po number": "po_number",
    "ship to": "ship_to",
    "qty": "quantity",
    "unit cost": "unit_price",
    "unit price": "unit_price",
    "ext price": "total_price",
    "total": "total_price",
}

_DATE_FORMATS = (
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%m/%d/%y",
    "%d-%b-%y",
    "%d-%b-%Y",
    "%B %d, %Y",
    "%b %d, %Y",
)


def _client() -> OpenAIClient:
    global _openai
    if _openai is None:
        _openai = OpenAIClient(OpenAISettings())
    return _openai


# ---------------------------------------------------------------------------
# Body helpers
# ---------------------------------------------------------------------------

def _strip_html(html: str) -> str:
    html = re.sub(r"(?is)<script[^>]*>.*?</script>", " ", html)
    html = re.sub(r"(?is)<style[^>]*>.*?</style>", " ", html)
    html = re.sub(r"<[^>]+>", " ", html)
    return unescape(html)


def _clean_body(body: str) -> str:
    if "<" in body and ">" in body:
        body = _strip_html(body)
    body = re.sub(r"[ \t]+\n", "\n", body)
    body = re.sub(r"\n{3,}", "\n\n", body).strip()
    return body


# ---------------------------------------------------------------------------
# Spreadsheet helpers
# ---------------------------------------------------------------------------

def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    s = str(h).strip().lower()
    return _HEADER_ALIASES.get(s, s.replace(" ", "_"))


def _is_sheet(att) -> bool:
    ct = (att.content_type or "").lower()
    fn = (att.filename or "").lower()
    if any(fn.endswith(x) for x in (".xlsx", ".xls", ".xlsm", ".csv")):
        return True
    return any(
        x in ct
        for x in ("spreadsheet", "excel", "csv", "officedocument.spreadsheetml")
    )


def _rows_from_xlsx(path: str, filename: str) -> list[dict]:
    out: list[dict] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue
            headers = [_norm_header(h) for h in header_row]
            data_rows: list[dict] = []
            for row in rows_iter:
                if row is None or all(v is None for v in row):
                    continue
                d = {
                    headers[i]: row[i]
                    for i in range(min(len(headers), len(row)))
                    if headers[i]
                }
                data_rows.append(d)
            out.append({"sheet_name": sheet_name, "rows": data_rows})
    finally:
        wb.close()
    return [{"filename": filename, "sheets": out}]


def _rows_from_csv(path: str, filename: str) -> list[dict]:
    df = pd.read_csv(path)
    df.columns = [_norm_header(c) for c in df.columns]
    records = df.to_dict(orient="records")
    return [
        {
            "filename": filename,
            "sheets": [{"sheet_name": "csv", "rows": records}],
        }
    ]


def _parse_sheets(attachments, errs: list[str]) -> str:
    """Parse all spreadsheet attachments and return a JSON text block."""
    blocks: list[dict] = []
    for att in attachments:
        if not _is_sheet(att):
            continue
        path = None
        try:
            raw = b64decode_bytes(att.data_base64)
            fn = att.filename.lower()
            if fn.endswith(".csv"):
                path = write_temp_bytes(".csv", raw)
                blocks.extend(_rows_from_csv(path, att.filename))
            else:
                ext = os.path.splitext(fn)[1] or ".xlsx"
                path = write_temp_bytes(ext, raw)
                blocks.extend(_rows_from_xlsx(path, att.filename))
        except Exception as e:
            logger.exception("extract_po excel_parser %s: %s", att.filename, e)
            errs.append(f"extract_po excel_parser {att.filename}: {e}")
        finally:
            unlink_silent(path)

    if not blocks:
        return ""

    parts: list[str] = []
    for block in blocks:
        fn = block.get("filename", "unknown")
        for sh in block.get("sheets") or []:
            name = sh.get("sheet_name", "")
            rows = sh.get("rows") or []
            parts.append(
                f'SPREADSHEET "{fn}" — Sheet "{name}":\n'
                + json.dumps(rows, default=str, indent=2)
            )
    return "\n\n".join(parts)


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

def _is_pdf(att) -> bool:
    ct = (att.content_type or "").lower()
    fn = (att.filename or "").lower()
    return "pdf" in ct or fn.endswith(".pdf")


def _render_pdf_pages(attachments, errs: list[str]) -> list[dict]:
    """Render every page of every PDF attachment as a base64 PNG image_url block."""
    image_blocks: list[dict] = []
    for att in attachments:
        if not _is_pdf(att):
            continue
        path = None
        try:
            import fitz

            raw = b64decode_bytes(att.data_base64)
            path = write_temp_bytes(".pdf", raw)
            doc = fitz.open(path)
            try:
                for i in range(doc.page_count):
                    page = doc.load_page(i)
                    pix = page.get_pixmap(dpi=150)
                    png = pix.tobytes("png")
                    b64 = base64.standard_b64encode(png).decode("ascii")
                    image_blocks.append({
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{b64}"},
                    })
            finally:
                doc.close()
        except Exception as e:
            logger.exception("extract_po pdf_render %s: %s", att.filename, e)
            errs.append(f"extract_po pdf_render {att.filename}: {e}")
        finally:
            unlink_silent(path)
    return image_blocks


# ---------------------------------------------------------------------------
# Normalisation helpers (deterministic, no LLM)
# ---------------------------------------------------------------------------

def _parse_date(s: str | None) -> str | None:
    if not s or not str(s).strip():
        return None
    s = str(s).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s


def _clean_money(s: float | int | str | None) -> float | None:
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    t = re.sub(r"[$,\s]", "", str(s))
    try:
        return float(t)
    except ValueError:
        return None


def _clean_customer(s: str | None) -> str | None:
    if not s:
        return None
    x = s.strip()
    for suf in (" Inc.", " LLC", " Corp.", " Corporation", " Ltd."):
        if x.endswith(suf):
            x = x[: -len(suf)].strip()
    return x.title() if x else None


def _norm_item(it: POItem) -> POItem:
    sku = (it.sku or "").strip().upper() or None
    return POItem(
        sku=sku,
        description=it.description.strip() if it.description else None,
        quantity=int(it.quantity) if it.quantity is not None else None,
        unit_price=_clean_money(it.unit_price),
        total_price=_clean_money(it.total_price),
        destination=it.destination.strip() if it.destination else None,
    )


def _normalize(po: ExtractedPO) -> ExtractedPO:
    items = [_norm_item(i) for i in (po.items or [])]
    return ExtractedPO(
        po_number=(po.po_number or "").strip() or None,
        customer=_clean_customer(po.customer),
        po_date=_parse_date(po.po_date),
        ship_date=_parse_date(po.ship_date),
        cancel_date=_parse_date(po.cancel_date),
        items=items,
        destinations=list(po.destinations or []),
        source_type=po.source_type,
        raw_confidence=po.raw_confidence,
        total_amount=_clean_money(po.total_amount),
        currency=(po.currency or "USD").upper(),
        payment_terms=po.payment_terms,
        ship_to=po.ship_to,
        bill_to=po.bill_to,
    )


# ---------------------------------------------------------------------------
# Main node
# ---------------------------------------------------------------------------

def extract_po_node(state: AgentState) -> dict:
    """Parse body + spreadsheets + PDFs, call GPT-4o vision, normalise result."""
    email = state["email"]
    errs = list(state["errors"])

    # --- Step 1: clean email body ---
    body_text = _clean_body(email.body or "")

    # --- Step 2: parse spreadsheet attachments to text ---
    sheet_text = _parse_sheets(email.attachments, errs)

    # --- Step 3: render PDF pages as images ---
    pdf_images = _render_pdf_pages(email.attachments, errs)

    # --- Step 4: build multi-modal message ---
    content_parts: list[dict] = []
    if body_text:
        content_parts.append({"type": "text", "text": f"EMAIL BODY:\n{body_text}"})
    if sheet_text:
        content_parts.append({"type": "text", "text": f"SPREADSHEET DATA:\n{sheet_text}"})
    content_parts.extend(pdf_images)

    if not content_parts:
        errs.append("extract_po: no usable content found (no body, PDFs, or spreadsheets)")
        return {"extracted_po": None, "normalized_po": None, "errors": errs}

    messages: list[dict[str, Any]] = [
        {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
        {"role": "user", "content": content_parts},
    ]

    # --- Step 5: call GPT-4o ---
    try:
        cli = _client()
        if not cli.enabled:
            errs.append("extract_po: OPENAI_API_KEY not set")
            return {"extracted_po": None, "normalized_po": None, "errors": errs}

        model = cli.settings.extraction_vision_model
        raw = cli.chat_completion(messages, model=model, json_mode=True)

        try:
            po = ExtractedPO.model_validate_json(raw)
        except Exception:
            retry_parts = list(content_parts) + [
                {"type": "text", "text": "\n\nOutput valid JSON only matching the schema. No markdown."},
            ]
            messages2: list[dict[str, Any]] = [
                {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                {"role": "user", "content": retry_parts},
            ]
            raw2 = cli.chat_completion(messages2, model=model, json_mode=True)
            po = ExtractedPO.model_validate_json(raw2)

    except Exception as e:
        logger.exception("extract_po LLM: %s", e)
        errs.append(f"Extraction failed: {e}")
        return {"extracted_po": None, "normalized_po": None, "errors": errs}

    # --- Step 6: deterministic normalisation ---
    try:
        normalized = _normalize(po)
    except Exception as e:
        logger.exception("extract_po normalizer: %s", e)
        errs.append(f"extract_po normalizer: {e}")
        normalized = po

    return {"extracted_po": po, "normalized_po": normalized, "errors": errs}
